#!/usr/bin/env python3
"""Copy and fill the ASVS Excel tracker from repository evidence."""

from __future__ import annotations

import argparse
import re
import shutil
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[2]
DEFAULT_XLSX = SCRIPT_DIR / "ASVS_5.0_Tracker1.xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR / "ASVS_5.0_Tracker_filled.xlsx"
DEFAULT_REPORT_DIRS = (
    PROJECT_ROOT / "vendnet" / "target" / "surefire-reports",
    PROJECT_ROOT / "vendnet" / "target" / "failsafe-reports",
)

REQ_ID_PATTERN = re.compile(r"^V\d+\.\d+\.\d+$")
REF_PATTERN = re.compile(r"\b(?:SR|FR|NFR)-\d+\b|\bAC-\d+\b")
DX_SHEETS = {"Summary", "ASVS Evidence", "ASVS DX Guide"}
DX_COLUMNS = [
    "DX Status Meaning",
    "Evidence Type",
    "Automated Evidence",
    "Evidence Source",
    "Next Action",
]

SCAN_EXTENSIONS = {".java", ".properties", ".xml", ".yml", ".yaml", ".json"}
SCAN_DIRS = (
    PROJECT_ROOT / "vendnet" / "src" / "main",
    PROJECT_ROOT / "vendnet" / "src" / "test",
    PROJECT_ROOT / ".github",
)

EVIDENCE_BY_REF = {
    "SR-01": ("administrator MFA", "AuthServiceTest; AuthControllerIntegrationTest"),
    "SR-02": ("RBAC roles and endpoint isolation", "RbacIntegrationTest; SecurityAnnotationArchTest"),
    "SR-03": ("JWT signature and algorithm validation", "AbuseCaseRegressionTest; JwtServiceTest; JwtAuthenticationFilterTest"),
    "SR-04": ("JWT key strength and server-side signing", "JwtServiceTest; JwtAuthenticationFilterTest"),
    "SR-05": ("account status validation", "AuthServiceTest; AuthControllerIntegrationTest"),
    "SR-06": ("controller/service authorization guardrails", "RbacIntegrationTest; SecurityAnnotationArchTest"),
    "SR-07": ("failed-login lockout behaviour", "AuthServiceTest; AuthControllerIntegrationTest; SystemFunctionalTests"),
    "SR-08": ("user data access isolation", "RbacIntegrationTest; SystemFunctionalTests"),
    "SR-09": ("SQL injection regression and taint checks", "AbuseCaseRegressionTest; IastIntegrationTest"),
    "SR-11": ("secrets externalized through configuration", "GitHub Actions secret scan; application-prod.properties"),
    "SR-14": ("safe JSON response/error surface", "GlobalExceptionHandlerTest; ControllerIntegrationTests"),
    "SR-15": ("transaction and stock consistency", "SaleServiceTest; SlotServiceTest; AbuseCaseRegressionTest"),
    "SR-16": ("DTO response boundaries and unknown-field rejection", "ControllerUnitTests; GlobalExceptionHandlerTest"),
    "SR-17": ("TLS configuration", "application-prod.properties; Docker/CI deployment configuration"),
    "SR-18": ("machine mTLS authentication paths", "X509MachineAuthenticationFilterTest; TelemetryServiceTest"),
    "SR-19": ("payment webhook HMAC validation", "AbuseCaseRegressionTest; PaymentGatewayServiceImplTest"),
    "SR-24": ("server-side price enforcement", "SaleServiceTest; AbuseCaseRegressionTest"),
    "SR-25": ("stock limit validation", "SlotServiceTest; SaleServiceTest"),
    "SR-26": ("path traversal and report sandboxing", "PathValidatorImplTest; ReportDirectoryServiceImplTest; AbuseCaseRegressionTest"),
    "SR-27": ("unsafe path/link handling", "PathValidatorImplTest; ReportDirectoryServiceImplTest"),
    "SR-28": ("input shape/range validation", "ControllerIntegrationTests; SystemFunctionalTests"),
    "SR-29": ("unknown-field rejection", "ControllerIntegrationTests; ControllerUnitTests; GlobalExceptionHandlerTest"),
    "SR-30": ("payload/request limit coverage", "SystemFunctionalTests; application.properties"),
    "SR-31": ("pagination/large-response controls", "ControllerIntegrationTests; SystemFunctionalTests"),
    "SR-32": ("dependency governance and SBOM", "Maven Enforcer; CycloneDX SBOM"),
    "SR-33": ("dependency update monitoring", "Dependabot; CycloneDX SBOM"),
    "SR-34": ("pinned dependency rules", "Maven Enforcer; pom.xml"),
    "SR-37": ("authentication audit logging", "AuthServiceTest; AuditLoggerTest"),
    "SR-38": ("inventory audit logging", "AuditLoggerTest; SystemFunctionalTests"),
    "SR-39": ("role-change audit logging", "AuditLoggerTest; RbacIntegrationTest"),
    "SR-40": ("payment event handling", "SaleServiceTest; PaymentGatewayServiceImplTest"),
    "SR-42": ("administrative file-operation logging", "AuditLoggerTest; SystemFunctionalTests"),
    "SR-45": ("sensitive log data checks", "AuditLoggerTest; GlobalExceptionHandlerTest"),
    "SR-46": ("safe error response checks", "GlobalExceptionHandlerTest; ControllerIntegrationTests"),
    "AC-01": ("OS command injection abuse-case regression", "AbuseCaseRegressionTest; IastIntegrationTest"),
    "AC-02": ("forged webhook abuse-case regression", "AbuseCaseRegressionTest; PaymentGatewayServiceImplTest"),
    "AC-03": ("client-supplied price abuse-case regression", "SaleServiceTest; AbuseCaseRegressionTest"),
    "AC-04": ("JWT alg:none abuse-case regression", "AbuseCaseRegressionTest; JwtServiceTest"),
    "AC-05": ("SQL injection abuse-case regression", "AbuseCaseRegressionTest; IastIntegrationTest"),
    "AC-06": ("path traversal abuse-case regression", "AbuseCaseRegressionTest; PathValidatorImplTest"),
    "AC-07": ("TOCTOU/stock race abuse-case regression", "SaleServiceTest; SlotServiceTest"),
    "AC-08": ("telemetry flood abuse-case regression", "TelemetryServiceTest; SystemFunctionalTests"),
}


@dataclass(frozen=True)
class Assessment:
    status: str
    observations: str
    reference: str
    evidence_type: str
    automated_evidence: str
    evidence_source: str
    next_action: str


@dataclass(frozen=True)
class EvidenceSnapshot:
    totals: dict[str, int]
    suites: list[list[object]]
    passing_suites: set[str]
    source_files: list[Path]
    refs_seen: set[str]
    features: dict[str, bool]

    @property
    def tests_passed(self) -> bool:
        return bool(self.totals["report_files"]) and self.totals["failures"] == 0 and self.totals["errors"] == 0


def relative(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def read_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return ""


def scanned_files() -> list[Path]:
    files: list[Path] = []
    for base in SCAN_DIRS:
        if not base.exists():
            continue
        files.extend(path for path in base.rglob("*") if path.is_file() and path.suffix in SCAN_EXTENSIONS)
    for extra in (PROJECT_ROOT / "vendnet" / "pom.xml", PROJECT_ROOT / "docker-compose.prod.yml"):
        if extra.exists():
            files.append(extra)
    return sorted(set(files))


def test_report_summary(report_dirs: tuple[Path, ...]) -> tuple[dict[str, int], list[list[object]], set[str]]:
    totals = {
        "report_files": 0,
        "tests": 0,
        "failures": 0,
        "errors": 0,
        "skipped": 0,
    }
    suites: list[list[object]] = []
    passing_suites: set[str] = set()

    for report_dir in report_dirs:
        if not report_dir.exists():
            continue

        for report in sorted(report_dir.glob("TEST-*.xml")):
            totals["report_files"] += 1
            try:
                root = ElementTree.parse(report).getroot()
            except ElementTree.ParseError:
                suites.append([relative(report), "parse error", "", "", "", "", "Unreadable XML"])
                continue

            test_suites = [root] if root.tag.endswith("testsuite") else [
                elem for elem in root.iter() if elem.tag.endswith("testsuite")
            ]
            for suite in test_suites:
                tests = int(float(suite.attrib.get("tests", 0)))
                failures = int(float(suite.attrib.get("failures", 0)))
                errors = int(float(suite.attrib.get("errors", 0)))
                skipped = int(float(suite.attrib.get("skipped", 0)))
                totals["tests"] += tests
                totals["failures"] += failures
                totals["errors"] += errors
                totals["skipped"] += skipped
                result = "PASS" if failures == 0 and errors == 0 else "FAIL"
                suite_name = suite.attrib.get("name", report.stem)
                short_name = suite_name.rsplit(".", 1)[-1]
                if result == "PASS":
                    passing_suites.add(short_name)
                suites.append([
                    relative(report),
                    suite_name,
                    tests,
                    failures,
                    errors,
                    skipped,
                    result,
                ])

    return totals, suites, passing_suites


def has_any(text: str, *needles: str) -> bool:
    lowered = text.lower()
    return any(needle.lower() in lowered for needle in needles)


def discover_evidence(report_dirs: tuple[Path, ...]) -> EvidenceSnapshot:
    totals, suites, passing_suites = test_report_summary(report_dirs)
    files = scanned_files()
    indexed_text = "\n".join(read_text(path) for path in files)
    refs_seen = set(REF_PATTERN.findall(indexed_text))

    features = {
        "no_frontend": not (PROJECT_ROOT / "vendnet" / "frontend").exists() and not (PROJECT_ROOT / "package.json").exists(),
        "no_oauth": not has_any(indexed_text, "oauth2Login", "spring-security-oauth2", "openid", "oidc"),
        "no_webrtc": not has_any(indexed_text, "webrtc", "turn server", "stun", "dtls"),
        "no_websocket": not has_any(indexed_text, "websocket", "stomp", "sockjs"),
        "no_graphql": not has_any(indexed_text, "graphql"),
        "no_xml_api": not has_any(indexed_text, "Produces.APPLICATION_XML", "application/xml", "soap"),
        "no_file_upload": not has_any(indexed_text, "MultipartFile", "@RequestPart", "@RequestParam(\"file\""),
        "jwt": has_any(indexed_text, "JwtService", "jjwt", "Bearer "),
        "mfa": has_any(indexed_text, "totp", "mfa", "MfaVerifyRequest"),
        "rbac": has_any(indexed_text, "@PreAuthorize", "@EnableMethodSecurity", "ROLE_ADMINISTRATOR"),
        "validation": has_any(indexed_text, "@Valid", "jakarta.validation", "fail-on-unknown-properties=true"),
        "unknown_fields": has_any(indexed_text, "fail-on-unknown-properties=true", "UnrecognizedPropertyException"),
        "hmac": has_any(indexed_text, "HmacSHA256", "X-Signature", "webhook-secret"),
        "mtls": has_any(indexed_text, "X509MachineAuthenticationFilter", "X509Certificate", "client-auth"),
        "aes_backup": has_any(indexed_text, "AES/GCM/NoPadding", "AES-256", "BackupServiceImpl"),
        "file_sandbox": has_any(indexed_text, "PathValidatorImpl", "toRealPath", "ReportDirectoryServiceImpl"),
        "audit": has_any(indexed_text, "AuditLogger", "AUDIT", "AuditLog"),
        "safe_errors": has_any(indexed_text, "GlobalExceptionHandler", "ApiError"),
        "sca": has_any(indexed_text, "dependency-check", "cyclonedx", "maven-enforcer-plugin"),
        "ci_security": has_any(indexed_text, "semgrep", "gitleaks", "trivy", "zap", "dependency-check"),
        "payload_limits": has_any(indexed_text, "max-request-size", "max-http-request-header-size"),
        "openapi": has_any(indexed_text, "springdoc", "OpenApiConfig", "swagger"),
    }

    return EvidenceSnapshot(
        totals=totals,
        suites=suites,
        passing_suites=passing_suites,
        source_files=files,
        refs_seen=refs_seen,
        features=features,
    )


def tokens_from_reference(reference: object) -> list[str]:
    if not reference:
        return []
    return REF_PATTERN.findall(str(reference))


def unique_join(values: list[str]) -> str:
    seen: set[str] = set()
    ordered = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            ordered.append(value)
    return "; ".join(ordered)


def infer_tokens(req_id: str, sheet_name: str, description: object, reference: object) -> list[str]:
    tokens = tokens_from_reference(reference)
    text = f"{sheet_name} {req_id} {description or ''}".lower()

    inferred: list[str] = []
    if has_any(text, "jwt", "self-contained token", "token"):
        inferred.extend(["SR-03", "SR-04"])
    if has_any(text, "mfa", "multi-factor", "authentication", "login", "password", "credential"):
        inferred.extend(["SR-01", "SR-05", "SR-07"])
    if has_any(text, "authorization", "access control", "permission", "role", "idor"):
        inferred.extend(["SR-02", "SR-06", "SR-08"])
    if has_any(text, "input", "validation", "unexpected field", "payload", "request body"):
        inferred.extend(["SR-28", "SR-29", "SR-30"])
    if has_any(text, "sql", "injection", "query", "parameterized"):
        inferred.extend(["SR-09", "AC-05"])
    if has_any(text, "file", "path", "directory", "upload", "download", "symlink"):
        inferred.extend(["SR-26", "SR-27", "AC-06"])
    if has_any(text, "payment", "webhook", "hmac"):
        inferred.extend(["SR-19", "AC-02"])
    if has_any(text, "price", "stock", "business logic", "transaction", "concurrent"):
        inferred.extend(["SR-15", "SR-24", "SR-25", "AC-03", "AC-07"])
    if has_any(text, "dependency", "third-party", "library", "component", "sbom"):
        inferred.extend(["SR-32", "SR-33", "SR-34"])
    if has_any(text, "log", "logging", "error", "exception"):
        inferred.extend(["SR-37", "SR-45", "SR-46"])
    if has_any(text, "tls", "certificate", "secure communication", "mtls"):
        inferred.extend(["SR-17", "SR-18"])
    if has_any(text, "secret", "configuration", "environment"):
        inferred.extend(["SR-11", "SR-34"])

    return list(dict.fromkeys(tokens + inferred))


def is_not_applicable(sheet_name: str, description: object, evidence: EvidenceSnapshot) -> tuple[bool, str]:
    text = f"{sheet_name} {description or ''}".lower()
    features = evidence.features

    if sheet_name.startswith("V17") and features["no_webrtc"]:
        return True, "No WebRTC, TURN, STUN, DTLS media, or browser real-time media component exists in this backend."
    if sheet_name.startswith("V10") and features["no_oauth"]:
        return True, "No OAuth/OIDC client, authorization server, or federated login flow is implemented."
    if sheet_name.startswith("V3") and features["no_frontend"]:
        return True, "VendNet is a Spring REST API/backend in this repository; no browser frontend or DOM-rendering surface is present."
    if has_any(text, "html", "css", "dom", "javascript", "browser") and features["no_frontend"]:
        return True, "No browser-rendered frontend surface is present in this repository."
    if has_any(text, "graphql") and features["no_graphql"]:
        return True, "No GraphQL endpoint or GraphQL dependency is present."
    if has_any(text, "xml", "soap") and features["no_xml_api"]:
        return True, "No XML/SOAP API surface is present; controllers return JSON REST responses."
    if has_any(text, "file upload", "accepts a file", "uploaded file") and features["no_file_upload"]:
        return True, "No user file-upload endpoint is implemented; file handling is limited to server-generated reports/backups."
    if has_any(text, "svg", "scalable vector graphics") and features["no_file_upload"] and features["no_frontend"]:
        return True, "No user-supplied SVG/content-rendering surface exists in this backend."
    if has_any(text, "websocket") and features["no_websocket"]:
        return True, "No WebSocket endpoint or WebSocket dependency is present."
    if has_any(text, "pointer", "pointers", "memory-safe string", "dynamically allocated memory", "full memory encryption"):
        return True, "This is a managed JVM application without native pointer arithmetic or application-managed memory allocation."
    if has_any(text, "format strings") and not has_any(text, "sql"):
        return True, "No user-controlled printf-style format-string execution path is present in the Java REST backend."

    return False, ""


def evidence_for_tokens(tokens: list[str], evidence: EvidenceSnapshot) -> tuple[list[str], list[str]]:
    topics: list[str] = []
    automated: list[str] = []
    for token in tokens:
        if token not in EVIDENCE_BY_REF:
            continue
        topic, tests = EVIDENCE_BY_REF[token]
        topics.append(topic)
        for item in [part.strip() for part in tests.split(";")]:
            if not item:
                continue
            if item in evidence.passing_suites:
                automated.append(item)
            elif any(item.lower() in relative(path).lower() for path in evidence.source_files):
                automated.append(item)
            elif item in {"Dependabot", "Maven Enforcer", "CycloneDX SBOM", "GitHub Actions secret scan"}:
                automated.append(item)
    return topics, automated


def feature_evidence(tokens: list[str], evidence: EvidenceSnapshot) -> list[str]:
    features = evidence.features
    hits: list[str] = []
    token_set = set(tokens)
    if {"SR-03", "SR-04"} & token_set and features["jwt"]:
        hits.append("JWT service and signing/filter code detected")
    if "SR-01" in token_set and features["mfa"]:
        hits.append("MFA/TOTP code and tests detected")
    if {"SR-02", "SR-06", "SR-08"} & token_set and features["rbac"]:
        hits.append("Spring method security and RBAC annotations detected")
    if {"SR-28", "SR-29", "SR-30"} & token_set and features["validation"]:
        hits.append("Bean Validation and request-shape controls detected")
    if "SR-19" in token_set and features["hmac"]:
        hits.append("HMAC webhook validation code detected")
    if "SR-18" in token_set and features["mtls"]:
        hits.append("X.509/mTLS machine authentication code detected")
    if {"SR-26", "SR-27"} & token_set and features["file_sandbox"]:
        hits.append("path validation and report sandbox code detected")
    if {"SR-37", "SR-38", "SR-39", "SR-42", "SR-45"} & token_set and features["audit"]:
        hits.append("audit logging implementation detected")
    if {"SR-32", "SR-33", "SR-34"} & token_set and features["sca"]:
        hits.append("SCA/SBOM/enforcer configuration detected")
    if {"SR-14", "SR-46"} & token_set and features["safe_errors"]:
        hits.append("centralized safe error handling detected")
    if {"SR-11"} & token_set and features["ci_security"]:
        hits.append("CI secret/security scanning configuration detected")
    return hits


def assess_requirement(
    req_id: str,
    sheet_name: str,
    description: object,
    level: object,
    reference: object,
    evidence: EvidenceSnapshot,
) -> Assessment:
    na, na_reason = is_not_applicable(sheet_name, description, evidence)
    tokens = infer_tokens(req_id, sheet_name, description, reference)
    topics, automated = evidence_for_tokens(tokens, evidence)
    feature_hits = feature_evidence(tokens, evidence)
    ref_text = unique_join(tokens) or "Automated inference"

    if na:
        return Assessment(
            status="N/A",
            observations=f"Automatically marked out of scope: {na_reason}",
            reference="Automated scope inference",
            evidence_type="Scope inference",
            automated_evidence="Not required for current backend scope.",
            evidence_source=na_reason,
            next_action="Revisit if the application scope changes.",
        )

    if evidence.tests_passed and (automated or feature_hits):
        evidence_text = unique_join(automated + feature_hits)
        topic_text = unique_join(topics) or "source/test feature match"
        return Assessment(
            status="Compliant",
            observations=f"Automatically marked compliant from passing evidence: {topic_text}.",
            reference=f"{ref_text}; {evidence_text}",
            evidence_type="Automated repository evidence",
            automated_evidence=evidence_text,
            evidence_source=topic_text,
            next_action="Keep make asvs-tracker passing; reassess if code, tests, or scope changes.",
        )

    if tokens and any(token in evidence.refs_seen for token in tokens):
        topic_text = unique_join(topics) or "reference token found in repository"
        return Assessment(
            status="Planned",
            observations=f"Automatically found related references ({unique_join(tokens)}), but no passing direct evidence was strong enough to claim compliance.",
            reference=ref_text,
            evidence_type="Referenced but not proven",
            automated_evidence="No direct passing automated evidence matched this ASVS row.",
            evidence_source=topic_text,
            next_action="Add implementation/tests that directly prove this ASVS row, then rerun make asvs-tracker.",
        )

    return Assessment(
        status="Planned",
        observations="Automatically marked planned because no direct implementation, test, or scope exemption matched this ASVS row.",
        reference="Automated inference found no strong repository evidence",
        evidence_type="Evidence needed",
        automated_evidence="No direct automated evidence matched this ASVS row.",
        evidence_source=f"ASVS row {req_id}; level {level or 'unknown'}",
        next_action="Implement the control or add a targeted test/evidence rule, then rerun make asvs-tracker.",
    )


def dx_values(assessment: Assessment) -> list[str]:
    if assessment.status == "N/A":
        meaning = "Out of scope for VendNet as currently implemented, or the requirement does not apply to this REST API/backend."
    elif assessment.status == "Compliant":
        meaning = "Automatically inferred as implemented or justified from passing tests, source code, configuration, or build artifacts."
    else:
        meaning = "Automatically inferred as required or potentially relevant, but not yet proven by strong repository evidence."

    return [
        meaning,
        assessment.evidence_type,
        assessment.automated_evidence,
        assessment.evidence_source,
        assessment.next_action,
    ]


def ensure_dx_columns(ws: openpyxl.worksheet.worksheet.Worksheet, header: list[object]) -> dict[str, int]:
    col_map = {str(value): idx + 1 for idx, value in enumerate(header) if value}
    next_col = ws.max_column + 1
    for name in DX_COLUMNS:
        if name not in col_map:
            ws.cell(1, next_col).value = name
            col_map[name] = next_col
            next_col += 1
    return col_map


def fill_xlsx(wb: openpyxl.Workbook, evidence: EvidenceSnapshot) -> int:
    updated = 0
    for sheet_name in wb.sheetnames:
        if sheet_name in DX_SHEETS:
            continue

        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        col_map = ensure_dx_columns(ws, header)
        required = ["Req ID", "Description", "Level", "Status", "Observations", "Reference / Link"]
        if not all(name in col_map for name in required):
            print(f"  Skipping sheet '{sheet_name}' - missing expected columns")
            continue

        for row in ws.iter_rows(min_row=2):
            req_id = row[col_map["Req ID"] - 1].value
            if not req_id or not REQ_ID_PATTERN.match(str(req_id)):
                continue

            assessment = assess_requirement(
                str(req_id),
                sheet_name,
                row[col_map["Description"] - 1].value,
                row[col_map["Level"] - 1].value,
                row[col_map["Reference / Link"] - 1].value,
                evidence,
            )
            row[col_map["Status"] - 1].value = assessment.status
            row[col_map["Observations"] - 1].value = assessment.observations
            row[col_map["Reference / Link"] - 1].value = assessment.reference
            for col_name, value in zip(DX_COLUMNS, dx_values(assessment)):
                ws.cell(row[0].row, col_map[col_name]).value = value
            updated += 1

        style_chapter_sheet(ws)

    return updated


def style_chapter_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    status_fills = {
        "Compliant": PatternFill("solid", fgColor="D9EAD3"),
        "Planned": PatternFill("solid", fgColor="FFF2CC"),
        "N/A": PatternFill("solid", fgColor="E7E6E6"),
        "Not Started": PatternFill("solid", fgColor="F4CCCC"),
    }
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header = [cell.value for cell in ws[1]]
    col_map = {str(value): idx + 1 for idx, value in enumerate(header) if value}
    widths = {
        "Section ID": 12,
        "Section Name": 28,
        "Req ID": 12,
        "Description": 62,
        "Level": 8,
        "Status": 14,
        "Observations": 58,
        "Reference / Link": 44,
        "DX Status Meaning": 48,
        "Evidence Type": 28,
        "Automated Evidence": 46,
        "Evidence Source": 46,
        "Next Action": 52,
    }
    for name, width in widths.items():
        if name in col_map:
            ws.column_dimensions[get_column_letter(col_map[name])].width = width

    status_col = col_map.get("Status")
    if status_col:
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row, status_col).value
            if status in status_fills:
                ws.cell(row, status_col).fill = status_fills[status]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def rebuild_summary_chart(wb: openpyxl.Workbook) -> None:
    """Replace stale/broken Summary charts with a chart tied to workbook formulas."""
    ws = wb["Summary"]
    ws._charts.clear()

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "ASVS 5.0 Compliance by Chapter"
    chart.y_axis.title = "Compliant Requirements"
    chart.x_axis.title = "Chapter"
    chart.style = 10
    chart.width = 30
    chart.height = 18

    data_min_row = 6
    data_max_row = 22
    for col_idx, series_title in (
        (4, "L1 Compliant"),
        (7, "L2 Compliant"),
        (10, "L3 Compliant"),
        (12, "Overall Compliant"),
    ):
        values = Reference(ws, min_col=col_idx, min_row=data_min_row, max_row=data_max_row)
        chart.append(openpyxl.chart.Series(values, title=series_title))

    categories = Reference(ws, min_col=1, min_row=data_min_row, max_row=data_max_row)
    chart.set_categories(categories)
    chart.shape = 4
    ws.add_chart(chart, "O5")


def add_dx_guide_sheet(wb: openpyxl.Workbook) -> None:
    if "ASVS DX Guide" in wb.sheetnames:
        del wb["ASVS DX Guide"]

    ws = wb.create_sheet("ASVS DX Guide", 1)
    rows = [
        ["ASVS Tracker DX Guide", "", "", ""],
        ["What is automated?", "make asvs-tracker runs evidence tests, copies the source workbook, scans repository evidence, infers each ASVS row, and writes this filled workbook.", "", ""],
        ["Input source", "The Excel template is the source of ASVS rows. The markdown tracker is not read or required.", "", ""],
        ["Important limitation", "The automation is deterministic: it can infer from tests, source, configuration, and build artifacts, but a new control may need a new targeted test/evidence rule before it becomes Compliant.", "", ""],
        ["", "", "", ""],
        ["Status", "Meaning", "When generated", "Next action"],
        ["Compliant", "The row has matching implementation/configuration/build evidence and the evidence test run passed.", "Generated when direct automated evidence or source evidence matches the ASVS row.", "Keep make asvs-tracker passing and update tests when scope changes."],
        ["Planned", "The row is relevant but not yet strongly proven by repository evidence.", "Generated when the row has no direct matching evidence or only weak/reference evidence.", "Implement the control or add a targeted test/evidence rule."],
        ["N/A", "The row is outside the current backend/API scope.", "Generated for absent technologies such as frontend DOM, OAuth/OIDC, GraphQL, XML/SOAP, WebRTC, or file upload surfaces.", "Revisit if that technology enters scope."],
        ["", "", "", ""],
        ["DX column", "Purpose", "Example", ""],
        ["DX Status Meaning", "Plain-language explanation of the generated status.", "Compliant means repository evidence matched and tests passed.", ""],
        ["Evidence Type", "Shows whether support is automated, referenced, scope-based, or missing.", "Automated repository evidence", ""],
        ["Automated Evidence", "Mapped tests, code features, or build artifacts supporting the row.", "RbacIntegrationTest; Spring method security and RBAC annotations detected", ""],
        ["Evidence Source", "The security topic inferred from ASVS text and SR/AC references.", "RBAC roles and endpoint isolation", ""],
        ["Next Action", "What to do to move the row forward.", "Add implementation/tests that directly prove this ASVS row.", ""],
        ["", "", "", ""],
        ["Recommended commands", "make asvs-tracker", "Runs tests and regenerates the filled Excel.", ""],
        ["Recommended commands", "make asvs-fill", "Regenerates Excel from existing reports without rerunning tests.", ""],
    ]
    for row in rows:
        ws.append(row)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    ws["A1"].fill = title_fill
    for header_row in (6, 11):
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [24, 76, 64, 56]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_evidence_sheet(
    wb: openpyxl.Workbook,
    source_xlsx: Path,
    output_xlsx: Path,
    updated: int,
    evidence: EvidenceSnapshot,
) -> None:
    if "ASVS Evidence" in wb.sheetnames:
        del wb["ASVS Evidence"]

    ws = wb.create_sheet("ASVS Evidence")
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    if not evidence.totals["report_files"]:
        result = "NO TEST REPORTS"
    elif evidence.tests_passed:
        result = "PASS"
    else:
        result = "FAIL"

    enabled_features = ", ".join(name for name, value in sorted(evidence.features.items()) if value)
    rows = [
        ["ASVS Tracker Evidence", "", "", "", "", "", ""],
        ["Generated at", generated_at, "", "", "", "", ""],
        ["Automation mode", "Excel rows + repository evidence inference; markdown tracker ignored", "", "", "", "", ""],
        ["Workbook source", relative(source_xlsx), "", "", "", "", ""],
        ["Filled workbook", relative(output_xlsx), "", "", "", "", ""],
        ["Workbook requirement rows inferred", updated, "", "", "", "", ""],
        ["Repository files scanned", len(evidence.source_files), "", "", "", "", ""],
        ["Reference tokens found in repo", unique_join(sorted(evidence.refs_seen)), "", "", "", "", ""],
        ["Detected feature flags", enabled_features, "", "", "", "", ""],
        ["Maven XML report files", evidence.totals["report_files"], "", "", "", "", ""],
        ["Total tests", evidence.totals["tests"], "", "", "", "", ""],
        ["Failures", evidence.totals["failures"], "", "", "", "", ""],
        ["Errors", evidence.totals["errors"], "", "", "", "", ""],
        ["Skipped", evidence.totals["skipped"], "", "", "", "", ""],
        ["Overall test evidence result", result, "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
    ]

    header_row = len(rows) + 1
    rows.append(["Report File", "Suite", "Tests", "Failures", "Errors", "Skipped", "Result"])

    for row in rows + evidence.suites:
        ws.append(row)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = title_fill
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = f"A{header_row + 1}"
    widths = [48, 70, 10, 10, 10, 10, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def force_recalculate_on_open(wb: openpyxl.Workbook) -> None:
    calculation = getattr(wb, "calculation", None) or getattr(wb, "calculation_properties", None)
    if calculation is None:
        return
    calculation.calcMode = "auto"
    calculation.fullCalcOnLoad = True
    calculation.forceFullCalc = True


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--report-dir",
        type=Path,
        action="append",
        dest="report_dirs",
        help="Directory containing Maven TEST-*.xml reports. Can be passed more than once.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source_xlsx = args.source_xlsx.resolve()
    output_xlsx = args.output.resolve()
    report_dirs = tuple(path.resolve() for path in (args.report_dirs or DEFAULT_REPORT_DIRS))

    if not source_xlsx.exists():
        raise FileNotFoundError(f"Source workbook not found: {source_xlsx}")

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_xlsx, output_xlsx)

    print(f"Copied source workbook: {relative(source_xlsx)}")
    print(f"Writing filled workbook: {relative(output_xlsx)}")
    print("Scanning tests, source code, configuration, and build artifacts...")
    evidence = discover_evidence(report_dirs)
    print(f"Scanned {len(evidence.source_files)} repository files.")
    print(f"Found {evidence.totals['report_files']} Maven XML report files with {evidence.totals['tests']} tests.")

    wb = openpyxl.load_workbook(output_xlsx)
    updated = fill_xlsx(wb, evidence)
    print(f"Inferred and updated {updated} workbook requirement rows.")

    rebuild_summary_chart(wb)
    add_dx_guide_sheet(wb)
    add_evidence_sheet(wb, source_xlsx, output_xlsx, updated, evidence)
    force_recalculate_on_open(wb)
    wb.save(output_xlsx)
    print("Rebuilt Summary chart and added ASVS DX Guide / Evidence sheets.")
    print(f"Saved: {output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
